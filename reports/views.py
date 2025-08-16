# reports/views.py
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.db.models import Sum
from django.http import HttpResponse
from django.utils.encoding import iri_to_uri
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from rest_framework import filters, permissions, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from .models import TransactionLedgerCombined
from .serializers import TransactionLedgerSerializer


Q2 = Decimal("0.01")  # 2-decimal quantize helper


class ReportPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = "page_size"
    max_page_size = 200


class TransactionLedgerViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Entity-Wise Report

    Required query params: start_date, end_date, entity
    Optional: cost_centre, transaction_type, source, min_amount, max_amount
    """
    permission_classes = [permissions.IsAuthenticated]
    queryset = TransactionLedgerCombined.objects.all().order_by("-date")
    serializer_class = TransactionLedgerSerializer
    pagination_class = ReportPagination

    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    # We’ll do date range ourselves; keep simple filters for FK/source:
    filterset_fields = ["entity", "cost_centre", "transaction_type", "source"]
    ordering_fields = [
        "date",
        "amount",
        "source",
        "cost_centre__name",
        "entity__name",
        "transaction_type__name",
    ]
    search_fields = ["remarks"]

    # ---- helpers -------------------------------------------------------------

    def _required_params(self):
        qp = self.request.query_params
        return qp.get("start_date"), qp.get("end_date"), qp.get("entity")

    def _parse_dates(self, start_date, end_date):
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d").date()
            ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            self._date_error = True
            return None, None
        if sd > ed:
            self._range_error = True
            return None, None
        return sd, ed

    def _parse_decimal(self, s):
        try:
            return Decimal(s)
        except (InvalidOperation, TypeError):
            return None

    # ---- queryset ------------------------------------------------------------

    def get_queryset(self):
        user = self.request.user
        qs = (
            super()
            .get_queryset()
            .select_related(
                "cost_centre",
                "entity",
                "transaction_type",
                "asset",
                "contract",
                "company",
            )
        )

        # Tenant scoping: user -> companies M2M
        if not getattr(user, "is_superuser", False):
            if not hasattr(user, "companies"):
                return qs.none()
            company_ids = list(user.companies.values_list("id", flat=True))
            if not company_ids:
                return qs.none()
            qs = qs.filter(company_id__in=company_ids)

        # Required filters (BRD)
        start_date, end_date, entity_id = self._required_params()
        if not start_date or not end_date or not entity_id:
            self._missing = {
                "start_date": bool(start_date),
                "end_date": bool(end_date),
                "entity": bool(entity_id),
            }
            return qs.none()

        sd, ed = self._parse_dates(start_date, end_date)
        if not sd or not ed:
            return qs.none()

        qs = qs.filter(date__gte=sd, date__lte=ed, entity_id=entity_id)

        # Optional amount range
        min_amount = self._parse_decimal(self.request.query_params.get("min_amount"))
        max_amount = self._parse_decimal(self.request.query_params.get("max_amount"))
        if min_amount is not None:
            qs = qs.filter(amount__gte=min_amount)
        if max_amount is not None:
            qs = qs.filter(amount__lte=max_amount)

        # Optional source
        source = self.request.query_params.get("source")
        if source in ("BANK", "CASH"):
            qs = qs.filter(source=source)

        return qs

    # ---- list with validation feedback --------------------------------------

    def list(self, request, *args, **kwargs):
        if getattr(self, "_date_error", False):
            return Response({"detail": "Invalid date format. Use YYYY-MM-DD."}, status=400)
        if getattr(self, "_range_error", False):
            return Response({"detail": "start_date cannot be after end_date."}, status=400)
        if getattr(self, "_missing", None):
            missing = [k for k, ok in self._missing.items() if not ok]
            return Response({"detail": f"Missing required filter(s): {', '.join(missing)}"}, status=400)
        return super().list(request, *args, **kwargs)

    # ---- summary -------------------------------------------------------------

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        if getattr(self, "_date_error", False):
            return Response({"detail": "Invalid date format. Use YYYY-MM-DD."}, status=400)
        if getattr(self, "_range_error", False):
            return Response({"detail": "start_date cannot be after end_date."}, status=400)
        if getattr(self, "_missing", None):
            missing = [k for k, ok in self._missing.items() if not ok]
            return Response({"detail": f"Missing required filter(s): {', '.join(missing)}"}, status=400)

        qs = self.filter_queryset(self.get_queryset())

        total_credit = qs.filter(amount__gt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        total_debit = qs.filter(amount__lt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        net = total_credit + total_debit  # debit is negative

        return Response(
            {
                "total_credit": total_credit.quantize(Q2, rounding=ROUND_HALF_UP),
                "total_debit": (-total_debit if total_debit < 0 else total_debit).quantize(Q2, rounding=ROUND_HALF_UP),
                "net": net.quantize(Q2, rounding=ROUND_HALF_UP),
            }
        )

    # ---- export --------------------------------------------------------------

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        if getattr(self, "_date_error", False):
            return Response({"detail": "Invalid date format. Use YYYY-MM-DD."}, status=400)
        if getattr(self, "_range_error", False):
            return Response({"detail": "start_date cannot be after end_date."}, status=400)
        if getattr(self, "_missing", None):
            missing = [k for k, ok in self._missing.items() if not ok]
            return Response({"detail": f"Missing required filter(s): {', '.join(missing)}"}, status=400)

        qs = self.filter_queryset(self.get_queryset())
        if not qs.exists():
            return Response({"detail": "No data to export."}, status=204)

        wb = Workbook()
        ws = wb.active
        ws.title = "Entity-Wise Report"

        headers = [
            "Date",
            "Source",
            "Amount",
            "Cost Centre",
            "Entity",
            "Transaction Type",
            "Asset",
            "Contract",
            "Remarks",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")

        for row in qs.iterator():
            ws.append([
                row.date,
                row.source,
                row.amount,
                getattr(row.cost_centre, "name", ""),
                getattr(row.entity, "name", ""),
                getattr(row.transaction_type, "name", ""),
                getattr(row.asset, "name", ""),
                getattr(row.contract, "name", ""),
                row.remarks or "",
            ])

        # formats
        for r in range(2, ws.max_row + 1):
            ws[f"A{r}"].number_format = numbers.FORMAT_DATE_DDMMMYYYY
            ws[f"C{r}"].number_format = numbers.FORMAT_NUMBER_00

        # summary
        ws.append([])
        total_credit = qs.filter(amount__gt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        total_debit = qs.filter(amount__lt=0).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        net = total_credit + total_debit

        for label, value in (
            ("Total Credit", total_credit),
            ("Total Debit", -total_debit if total_debit < 0 else total_debit),
            ("Net Amount", net),
        ):
            ws.append([label, "", value])
            ws[f"C{ws.max_row}"].number_format = numbers.FORMAT_NUMBER_00
            ws[f"A{ws.max_row}"].font = Font(bold=True)

        # simple best-fit widths
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in ws[letter])
            ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)

        entity_name = getattr(qs.first().entity, "name", "entity")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        fname = f"entity_wise_{entity_name}_{start_date}_to_{end_date}.xlsx"
        fname = iri_to_uri(fname.replace(" ", "_"))

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        wb.save(resp)
        return resp
